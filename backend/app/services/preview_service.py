from typing import Dict, Any, List
import pandas as pd
import json
from io import BytesIO
import base64
from datetime import datetime
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

class PreviewService:
    @staticmethod
    def generate_sample_data(columns: List[str], rows: int = 5) -> List[Dict[str, Any]]:
        """Generate sample data for preview"""
        sample_data = []
        for i in range(rows):
            row = {}
            for col in columns:
                # Generate appropriate sample data based on column name
                if 'date' in col.lower():
                    row[col] = (datetime.now()).strftime('%Y-%m-%d')
                elif any(metric in col.lower() for metric in ['score', 'rate', 'percentage']):
                    row[col] = round(0.75 + (i * 0.05), 2)
                elif 'count' in col.lower():
                    row[col] = 100 + (i * 10)
                else:
                    row[col] = f"Sample {col} {i+1}"
            sample_data.append(row)
        return sample_data

    def generate_csv_preview(self, template: Dict[str, Any]) -> str:
        """Generate CSV preview"""
        sample_data = self.generate_sample_data(template.get('columns', []))
        df = pd.DataFrame(sample_data)
        
        # Generate CSV
        csv_buffer = BytesIO()
        df.to_csv(csv_buffer, index=False)
        csv_content = csv_buffer.getvalue().decode('utf-8')
        
        # Return first few lines
        preview_lines = csv_content.split('\n')[:6]  # Header + 5 rows
        return '\n'.join(preview_lines)

    def generate_excel_preview(self, template: Dict[str, Any]) -> str:
        """Generate Excel preview and return as base64 image"""
        sample_data = self.generate_sample_data(template.get('columns', []))
        
        wb = Workbook()
        ws = wb.active
        
        # Apply styling from template
        styling = template.get('styling', {})
        header_fill = PatternFill(
            start_color=styling.get('headerColor', '4B0082').replace('#', ''),
            end_color=styling.get('headerColor', '4B0082').replace('#', ''),
            fill_type='solid'
        )
        header_font = Font(
            name=styling.get('fontFamily', 'Arial'),
            size=styling.get('fontSize', 11),
            bold=True,
            color='FFFFFF'
        )
        
        # Write headers
        for col_idx, column in enumerate(template.get('columns', []), 1):
            cell = ws.cell(row=1, column=col_idx, value=column)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Write sample data
        for row_idx, row_data in enumerate(sample_data, 2):
            for col_idx, value in enumerate(row_data.values(), 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = Font(
                    name=styling.get('fontFamily', 'Arial'),
                    size=styling.get('fontSize', 11)
                )
        
        # Save to buffer
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        # Convert to base64
        return base64.b64encode(excel_buffer.getvalue()).decode('utf-8')

    def generate_pdf_preview(self, template: Dict[str, Any]) -> str:
        """Generate PDF preview and return as base64"""
        sample_data = self.generate_sample_data(template.get('columns', []))
        
        # Create PDF buffer
        pdf_buffer = BytesIO()
        doc = SimpleDocTemplate(
            pdf_buffer,
            pagesize=letter,
            rightMargin=72,
            leftMargin=72,
            topMargin=72,
            bottomMargin=72
        )
        
        # Get styling from template
        styling = template.get('styling', {})
        
        # Create table data
        table_data = [template.get('columns', [])]  # Headers
        for row in sample_data:
            table_data.append([row[col] for col in template.get('columns', [])])
        
        # Create table with styling
        table = Table(table_data)
        style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(styling.get('headerColor', '#4B0082'))),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), styling.get('fontFamily', 'Helvetica-Bold')),
            ('FONTSIZE', (0, 0), (-1, 0), styling.get('fontSize', 12)),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTNAME', (0, 1), (-1, -1), styling.get('fontFamily', 'Helvetica')),
            ('FONTSIZE', (0, 1), (-1, -1), styling.get('fontSize', 10)),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ])
        table.setStyle(style)
        
        # Build PDF
        doc.build([table])
        
        # Convert to base64
        pdf_buffer.seek(0)
        return base64.b64encode(pdf_buffer.getvalue()).decode('utf-8')

    def generate_preview(self, template: Dict[str, Any]) -> Dict[str, Any]:
        """Generate preview based on template format"""
        preview_format = template.get('format', 'pdf')
        
        try:
            if preview_format == 'csv':
                preview_content = self.generate_csv_preview(template)
                preview_type = 'text'
            elif preview_format == 'excel':
                preview_content = self.generate_excel_preview(template)
                preview_type = 'excel'
            else:  # pdf
                preview_content = self.generate_pdf_preview(template)
                preview_type = 'pdf'
            
            return {
                'type': preview_type,
                'content': preview_content,
                'format': preview_format
            }
        except Exception as e:
            return {
                'type': 'error',
                'content': str(e),
                'format': preview_format
            }